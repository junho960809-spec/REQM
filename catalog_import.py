from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_HEADERS = {"품목코드", "품목명", "바코드"}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_item_catalog(path: str | Path) -> list[dict[str, Any]]:
    """Read the first sheet containing the required item catalog headers."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            header_row = None
            columns: dict[str, int] = {}
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True),
                start=1,
            ):
                values = {_text(value): index for index, value in enumerate(row) if _text(value)}
                if REQUIRED_HEADERS.issubset(values):
                    header_row, columns = row_number, values
                    break
            if header_row is None:
                continue

            records: list[dict[str, Any]] = []
            category = ""
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if "구분" in columns and columns["구분"] < len(row):
                    category = _text(row[columns["구분"]]) or category
                code = _text(row[columns["품목코드"]]) if columns["품목코드"] < len(row) else ""
                name = _text(row[columns["품목명"]]) if columns["품목명"] < len(row) else ""
                barcode = _text(row[columns["바코드"]]) if columns["바코드"] < len(row) else ""
                box_quantity = ""
                if "내품 수량" in columns and columns["내품 수량"] < len(row):
                    box_quantity = _text(row[columns["내품 수량"]])
                if not code and not name and not barcode:
                    continue
                if not code or not name:
                    raise ValueError(f"{sheet.title} {row[0] if row else ''}: 품목코드 또는 품목명이 비어 있습니다.")
                records.append(
                    {
                        "item_code": code,
                        "standard_name": name,
                        "barcode": barcode,
                        "box_quantity": box_quantity,
                        "category": category,
                    }
                )

            code_counts = Counter(row["item_code"].casefold() for row in records)
            barcode_counts = Counter(row["barcode"] for row in records if row["barcode"])
            duplicate_codes = [key for key, count in code_counts.items() if count > 1]
            duplicate_barcodes = [key for key, count in barcode_counts.items() if count > 1]
            if duplicate_codes or duplicate_barcodes:
                details = []
                if duplicate_codes:
                    details.append(f"중복 품목코드 {len(duplicate_codes)}개")
                if duplicate_barcodes:
                    details.append(f"중복 바코드 {len(duplicate_barcodes)}개")
                raise ValueError("파일을 가져올 수 없습니다: " + ", ".join(details))
            return records
    finally:
        workbook.close()
    raise ValueError("'품목코드', '품목명', '바코드' 열이 있는 시트를 찾지 못했습니다.")


def compare_catalog(
    records: list[dict[str, Any]],
    db_items: list[dict[str, Any]],
    db_barcodes: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    items_by_code = {str(row.get("item_code", "")).casefold(): row for row in db_items}
    barcodes_by_value = {
        str(row.get("barcode", "")): row
        for row in db_barcodes
        if str(row.get("barcode", ""))
    }
    new_items, existing_items, renamed_items = [], [], []
    new_barcodes, existing_barcodes, barcode_conflicts = [], [], []

    for record in records:
        existing_item = items_by_code.get(record["item_code"].casefold())
        if existing_item is None:
            new_items.append(record)
        else:
            existing_items.append(record)
            if str(existing_item.get("standard_name", "")).strip() != record["standard_name"]:
                renamed_items.append(record)

        barcode = record["barcode"]
        if not barcode:
            continue
        existing_barcode = barcodes_by_value.get(barcode)
        if existing_barcode is None:
            new_barcodes.append(record)
        elif str(existing_barcode.get("item_code", "")).casefold() == record["item_code"].casefold():
            existing_barcodes.append(record)
        else:
            barcode_conflicts.append(
                {**record, "db_item_code": str(existing_barcode.get("item_code", ""))}
            )

    return {
        "new_items": new_items,
        "existing_items": existing_items,
        "renamed_items": renamed_items,
        "new_barcodes": new_barcodes,
        "existing_barcodes": existing_barcodes,
        "barcode_conflicts": barcode_conflicts,
    }
