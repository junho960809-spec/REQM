import json
import re
import urllib.error
import urllib.request
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook


CODE_HEADERS = {"품목코드", "상품코드", "제품코드", "재고코드", "itemcode", "prodcd", "sku"}
NAME_HEADERS = {"품명", "품목명", "상품명", "제품명", "재고명", "재고매칭", "itemname", "proddes"}
QTY_HEADERS = {"수량", "총입고수량", "입고수량", "출고수량", "주문수량", "이동수량", "qty", "quantity"}


def compact(value: Any) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").lower())


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _excel_rows(path: str) -> list[list[str]]:
    sheets = []
    if Path(path).suffix.lower() == ".xls":
        book = xlrd.open_workbook(path, on_demand=True)
        for sheet in book.sheets():
            sheets.append([[_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)] for r in range(sheet.nrows)])
    else:
        book = load_workbook(path, read_only=True, data_only=True)
        for sheet in book.worksheets:
            sheets.append([[_text(value) for value in row] for row in sheet.iter_rows(values_only=True)])
    return [row for sheet in sheets for row in sheet]


def _pdf_rows(path: str) -> list[list[str]]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("PDF 분석 모듈이 없습니다. 최신 프로그램을 사용해 주세요.") from exc
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                rows.extend([[_text(cell) for cell in row] for row in table if row])
            if not tables:
                for line in (page.extract_text() or "").splitlines():
                    cells = [part.strip() for part in re.split(r"\s{2,}|\t", line) if part.strip()]
                    if cells:
                        rows.append(cells)
    return rows


def _find_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    best = None
    for row_index, row in enumerate(rows[:120]):
        found = {}
        for col_index, value in enumerate(row[:120]):
            key = compact(value)
            if key in CODE_HEADERS and "code" not in found:
                found["code"] = col_index
            if key in NAME_HEADERS and "name" not in found:
                found["name"] = col_index
            if key in QTY_HEADERS and "qty" not in found:
                found["qty"] = col_index
        score = len(found) + (2 if "qty" in found else 0)
        if ("code" in found or "name" in found) and "qty" in found and (best is None or score > best[0]):
            best = (score, row_index, found)
    if best is None:
        raise ValueError("품목코드/품목명과 수량 열을 찾지 못했습니다. 열 제목을 확인해 주세요.")
    return best[1], best[2]


def read_transfer_file(path: str) -> list[dict[str, str]]:
    suffix = Path(path).suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        rows = _excel_rows(path)
    elif suffix == ".pdf":
        rows = _pdf_rows(path)
    else:
        raise ValueError("지원 파일은 .xlsx, .xls, .pdf입니다.")
    header_row, columns = _find_header(rows)
    result = []
    for source_row, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        code_index, name_index, qty_index = columns.get("code"), columns.get("name"), columns["qty"]
        code = row[code_index].strip() if code_index is not None and code_index < len(row) else ""
        name = row[name_index].strip() if name_index is not None and name_index < len(row) else ""
        qty_text = row[qty_index].strip() if qty_index < len(row) else ""
        if not code and not name:
            continue
        try:
            quantity = float(qty_text.replace(",", ""))
        except ValueError:
            continue
        if quantity > 0:
            result.append({"source_row": str(source_row), "source_code": code, "source_name": name, "quantity": quantity})
    if not result:
        raise ValueError("파일에서 유효한 품목 행을 찾지 못했습니다.")
    return result


def _matched(raw: dict, item: dict, quantity: float, status: str, reason: str) -> dict:
    return {**raw, "item_code": str(item.get("item_code", "")), "item_name": str(item.get("standard_name", "")), "quantity": quantity, "status": status, "reason": reason}


def match_transfer_rows(raw_rows: list[dict], items: list[dict], products: list[dict], components: list[dict]) -> list[dict]:
    active_items = [row for row in items if row.get("is_active", True)]
    by_code = {str(row.get("item_code", "")).strip(): row for row in active_items}
    by_name = defaultdict(list)
    for row in active_items:
        by_name[compact(row.get("standard_name", ""))].append(row)
    product_by_name = defaultdict(list)
    for row in products:
        if row.get("is_active", True):
            product_by_name[compact(row.get("normalized_name") or row.get("original_name", ""))].append(row)
    component_by_product = defaultdict(list)
    for row in components:
        component_by_product[str(row.get("registered_product_id", ""))].append(row)

    result = []
    for raw in raw_rows:
        source_code, source_name, base_qty = raw["source_code"], raw["source_name"], float(raw["quantity"])
        if source_code in by_code:
            result.append(_matched(raw, by_code[source_code], base_qty, "exact", "품목코드 정확 일치"))
            continue
        exact_items = by_name.get(compact(source_name), []) if source_name else []
        if len(exact_items) == 1:
            result.append(_matched(raw, exact_items[0], base_qty, "exact" if not source_code else "similar", "품목명 정확 일치" if not source_code else "입력 코드는 미등록이나 품목명으로 확인"))
            continue
        exact_products = product_by_name.get(compact(source_name), []) if source_name else []
        if len(exact_products) == 1:
            product_components = component_by_product.get(str(exact_products[0].get("registered_product_id", "")), [])
            if product_components:
                for component in product_components:
                    result.append(_matched(raw, by_code.get(str(component.get("item_code", "")), {}), base_qty * float(component.get("quantity", 1)), "exact", "세트상품 구성품 자동 분해"))
                continue
        candidates = []
        source_key = compact(source_name)
        for item in active_items:
            item_key = compact(item.get("standard_name", ""))
            score = SequenceMatcher(None, source_key, item_key).ratio() if source_key else 0
            if item_key and item_key in source_key:
                score = max(score, 0.94)
            model_key = compact(item.get("model", ""))
            color_key = compact(item.get("color", ""))
            if model_key and model_key in source_key:
                score = max(score, 0.72)
                if color_key and color_key in source_key:
                    score = max(score, 0.90)
            source_has_case = "케이스" in source_key or "실리콘" in source_key
            item_has_case = "케이스" in item_key or "실리콘" in item_key
            if source_has_case == item_has_case:
                score += 0.08
            else:
                score -= 0.14
            source_handy = "핸디" in source_key
            item_handy = "핸디" in item_key or compact(item.get("form", "")) == "핸디형"
            if source_handy and item_handy:
                score += 0.06
            elif source_handy != item_handy:
                score -= 0.05
            if score >= 0.62:
                candidates.append((score, item))
        candidates.sort(key=lambda value: value[0], reverse=True)
        if candidates and (len(candidates) == 1 or candidates[0][0] - candidates[1][0] > 0.04):
            display_score = min(candidates[0][0], 0.99)
            result.append(_matched(raw, candidates[0][1], base_qty, "similar", f"품목명 유사 일치 {display_score:.0%}"))
        else:
            result.append({**raw, "item_code": "", "item_name": source_name, "status": "ambiguous" if candidates else "missing", "reason": "유사 후보가 여러 개라 확인 필요" if candidates else "DB에서 품목을 찾지 못함"})
    return result


def aggregate_transfer_rows(rows: list[dict]) -> list[dict]:
    aggregated, unresolved = {}, []
    for row in rows:
        code = str(row.get("item_code", "")).strip()
        if not code:
            unresolved.append(dict(row))
            continue
        if code not in aggregated:
            aggregated[code] = dict(row)
            aggregated[code]["quantity"] = 0.0
        aggregated[code]["quantity"] += float(row.get("quantity", 0))
        if row.get("status") != "exact":
            aggregated[code]["status"], aggregated[code]["reason"] = row.get("status"), row.get("reason")
    return list(aggregated.values()) + unresolved


class EcountClient:
    def __init__(self, com_code: str, user_id: str, api_cert_key: str, zone: str = "AB"):
        self.com_code, self.user_id, self.api_cert_key, self.zone = com_code, user_id, api_cert_key, zone

    @staticmethod
    def _post(url: str, payload: dict) -> dict:
        request = urllib.request.Request(url, data=json.dumps(payload, ensure_ascii=False).encode("utf-8"), headers={"Content-Type": "application/json"}, method="POST")
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            raise RuntimeError(f"이카운트 HTTP 오류 {exc.code}: {exc.read().decode('utf-8', errors='replace')}") from exc

    @staticmethod
    def _find_session(value: Any) -> str:
        if isinstance(value, dict):
            for key, child in value.items():
                if str(key).upper() == "SESSION_ID" and child:
                    return str(child)
                found = EcountClient._find_session(child)
                if found:
                    return found
        elif isinstance(value, list):
            for child in value:
                found = EcountClient._find_session(child)
                if found:
                    return found
        return ""

    def login(self) -> str:
        result = self._post(f"https://oapi{self.zone}.ecount.com/OAPI/V2/OAPILogin", {"COM_CODE": self.com_code, "USER_ID": self.user_id, "API_CERT_KEY": self.api_cert_key, "LAN_TYPE": "ko-KR", "ZONE": self.zone})
        session_id = self._find_session(result)
        if not session_id:
            raise RuntimeError(f"이카운트 로그인 실패: {result.get('Error') or result.get('Errors') or result}")
        return session_id

    def save_location_transfer(self, session_id: str, rows: list[dict], io_date: str, employee_code: str, from_code: str, to_code: str, remarks: str = "") -> dict:
        payload = {"LocationTranList": []}
        for index, row in enumerate(rows, start=1):
            quantity = float(row["quantity"])
            qty_text = str(int(quantity)) if quantity.is_integer() else str(quantity)
            payload["LocationTranList"].append({"BulkDatas": {"IO_DATE": io_date, "UPLOAD_SER_NO": str(index), "EMP_CD": employee_code, "WH_CD_F": from_code, "WH_CD_T": to_code, "PROD_CD": str(row["item_code"]), "PROD_DES": str(row.get("item_name", "")), "QTY": qty_text, "REMARKS": remarks}})
        return self._post(f"https://oapi{self.zone}.ecount.com/OAPI/V2/Others/SaveLocationTran?SESSION_ID={session_id}", payload)
