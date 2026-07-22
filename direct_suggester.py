from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from matcher import compact


IGNORE_OPTIONS = {"", "선택안함", "선택없음", "없음", "해당없음"}
ACCESSORY_WORDS = ("case", "케이스", "pouch", "파우치", "스티커", "sticker", "링", "strap", "스트랩", "패드")
COLOR_ALIASES = {
    "라밴더": "라벤더", "캐럿": "캐롯", "세이지": "세이지민트",
    "토마토": "토마토레드", "오렌지": "소프트오렌지",
}
COLORS = (
    "블랙", "화이트", "핑크", "그레이", "솔리드그레이", "그린", "블루", "레드",
    "토마토레드", "옐로우", "오렌지", "소프트오렌지", "퍼플", "베이지", "민트",
    "세이지민트", "라벤더", "버터", "캐롯", "샌드", "코발트블루", "에보니",
    "원더랜드", "모닝브리즈", "선셋", "올리브",
)


def _reference_path() -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / "assets" / "direct_conversion_reference.xlsx"


def _load_conversion_reference() -> dict[tuple[str, str], dict[str, str]]:
    path = _reference_path()
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["데이터_new"]
        result: dict[tuple[str, str], dict[str, str]] = {}
        for product, option, model, description, *_ in sheet.iter_rows(min_row=2, values_only=True):
            product_text = str(product or "").strip()
            if not product_text:
                continue
            key = (compact(product_text), compact(str(option or "")))
            # 중복 키는 먼저 등록된 값을 유지하여 잘못된 자동 덮어쓰기를 막는다.
            result.setdefault(
                key,
                {
                    "model": str(model or "").strip(),
                    "description": str(description or "").strip(),
                },
            )
        return result
    finally:
        workbook.close()


CONVERSION_REFERENCE = _load_conversion_reference()


def _clean_option(value: str) -> str:
    value = re.sub(r"^\s*\d+\s*[.)]\s*", "", value or "")
    value = re.sub(r"\bnew\b", "", value, flags=re.IGNORECASE)
    return " ".join(value.split()).strip()


def _parts(options: str) -> list[str]:
    parts = [_clean_option(part) for part in re.split(r"\s*(?:/|\+)\s*", options or "")]
    return [part for part in parts if compact(part) not in {compact(value) for value in IGNORE_OPTIONS}]


def _item_text(item: dict[str, Any]) -> str:
    return " ".join(str(item.get(key, "") or "") for key in ("item_code", "standard_name", "model", "color", "form"))


def _is_accessory(item: dict[str, Any]) -> bool:
    text = _item_text(item).lower()
    return any(word in text for word in ACCESSORY_WORDS)


def _model_keys(model: str) -> list[str]:
    key = compact(model)
    keys = [key] if key else []
    shortened = re.sub(r"[a-z]+$", "", key)
    if len(shortened) >= 5 and shortened not in keys:
        keys.append(shortened)
    return keys


def _option_terms(part: str) -> set[str]:
    normalized = part
    for old, new in COLOR_ALIASES.items():
        normalized = normalized.replace(old, new)
    terms = {color for color in COLORS if color in normalized}
    compact_part = compact(normalized)
    if compact_part:
        terms.add(compact_part)
    return {compact(term) for term in terms if compact(term)}


def _candidate_score(item: dict[str, Any], model_keys: list[str], part: str, accessory: bool) -> int:
    text = compact(_item_text(item))
    score = 0
    if model_keys:
        matches = [key for key in model_keys if key and key in text]
        if matches:
            score += 6 + max(len(key) for key in matches) // 3
        elif not accessory:
            return -100
    terms = _option_terms(part)
    color_terms = {compact(COLOR_ALIASES.get(color, color)) for color in COLORS if color in part}
    if color_terms:
        if any(term in text for term in color_terms):
            score += 6
        else:
            return -100
    non_color_terms = [term for term in terms if term not in color_terms and len(term) >= 3]
    if any(term in text for term in non_color_terms):
        score += 5
    item_accessory = _is_accessory(item)
    score += 4 if item_accessory == accessory else -5
    return score


def _choose(items: list[dict[str, Any]], model_keys: list[str], part: str, accessory: bool) -> tuple[dict[str, Any] | None, str]:
    scored = sorted(
        ((_candidate_score(item, model_keys, part, accessory), item) for item in items if item.get("is_active", True)),
        key=lambda pair: pair[0],
        reverse=True,
    )
    scored = [pair for pair in scored if pair[0] >= 6]
    if not scored:
        return None, "후보 없음"
    top_score, top_item = scored[0]
    if len(scored) > 1 and top_score - scored[1][0] < 3:
        names = ", ".join(str(item.get("item_code", "")) for _, item in scored[:3])
        return None, f"후보 여러 개: {names}"
    return top_item, f"추천 점수 {top_score}"


def suggest_direct_order(order: dict[str, Any], items: list[dict[str, Any]]) -> dict[str, Any]:
    source_product = str(order.get("product_name", "") or "").strip()
    source_options = str(order.get("options", "") or "").strip()
    reference = CONVERSION_REFERENCE.get((compact(source_product), compact(source_options)))
    model = str(order.get("model", "") or "").strip()
    option_text = source_options
    reference_note = ""
    reference_conflict = False
    if reference:
        model = reference.get("model") or model
        option_text = reference.get("description") or source_options
        # 참고표의 '크레용 케이블' 명칭을 실제 DB 모델 코드로 바꿔 길이별 품목을 구분한다.
        if "크레용" in compact(model):
            if "30cm" in compact(option_text):
                model = "CC30S"
            elif "60cm" in compact(option_text):
                model = "CC60S"
        # ACONE 애플워치 전용 구성에는 충전기 본품과 실리콘패드가 함께 출고된다.
        if compact(model) == "acone" and "애플워치" in source_options and "실리콘패드" not in option_text:
            option_text = f"{option_text} / 실리콘패드"
        source_colors = _option_terms(source_options) & {compact(COLOR_ALIASES.get(color, color)) for color in COLORS}
        mapped_colors = _option_terms(option_text) & {compact(COLOR_ALIASES.get(color, color)) for color in COLORS}
        reference_conflict = bool(source_colors and mapped_colors and not source_colors.issubset(mapped_colors))
        reference_note = "변환 참고표 적용"
        if reference_conflict:
            reference_note += " · 원본 옵션과 참고표 색상 충돌"
    option_parts = _parts(option_text)
    model_keys = _model_keys(model)
    selected: list[dict[str, Any]] = []
    notes: list[str] = []

    cable_set = "케이블" in compact(model) and len(option_parts) > 1
    base_part = option_parts[0] if option_parts else ""
    base, note = _choose(items, model_keys, base_part, accessory=False)
    notes.append(f"본품: {note}")
    if base:
        selected.append(base)
    for part in option_parts[1:]:
        component, note = _choose(items, model_keys, part, accessory=not cable_set)
        label = "세트품목" if cable_set else "추가옵션"
        notes.append(f"{label} {part}: {note}")
        if component:
            selected.append(component)

    expected = 1 + max(0, len(option_parts) - 1)
    high_confidence = bool(model_keys) and len(selected) == expected and not reference_conflict
    if reference_note:
        notes.insert(0, reference_note)
    if not high_confidence:
        return {"status": "review", "components": selected, "reason": " | ".join(notes)}
    return {"status": "auto", "components": selected, "reason": " | ".join(notes)}


def components_text(components: list[dict[str, Any]]) -> str:
    return " + ".join(f"{item.get('item_code', '')}×1" for item in components)


def component_payload(components: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "item_code": item.get("item_code", ""),
            "standard_name": item.get("standard_name", ""),
            "quantity": 1,
        }
        for item in components
    ]
