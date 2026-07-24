from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ecount_sales_core import (
    ReferenceCatalog,
    SmartStoreOrder,
    convert_orders,
    read_smartstore_orders,
    write_ecount_workbook,
)


class SalesCoreTests(unittest.TestCase):
    def setUp(self) -> None:
        self.catalog = ReferenceCatalog(
            items=[
                {"item_code": "MAIN", "representative_name": "본품"},
                {"item_code": "OPTION", "representative_name": "옵션품목"},
                {"item_code": "택배운송비", "representative_name": "배송비"},
            ],
            channels=[
                {"source_name": "리큐엠_스마트스토어", "ecount_customer_code": "AC008712", "ecount_customer_name": "샵N", "is_active": True}
            ],
            mappings=[
                {"mapping_key": "set-1", "source_channel": "리큐엠_스마트스토어", "normalized_source": "상품세트옵션핑크", "mapping_type": "set", "review_status": "confirmed", "is_active": True}
            ],
            mapping_components=[
                {"mapping_key": "set-1", "sequence": 1, "item_code": "MAIN", "quantity": 1},
                {"mapping_key": "set-1", "sequence": 2, "item_code": "OPTION", "quantity": 1},
            ],
            price_rules=[
                {"price_rule_key": "price-1", "source_channel": "리큐엠_스마트스토어", "source_product_name": "상품 세트", "source_options": "옵션 핑크", "normalized_source": "상품세트옵션핑크", "total_unit_price": 35400, "review_status": "confirmed", "is_active": True}
            ],
            price_components=[
                {"price_rule_key": "price-1", "sequence": 1, "item_code": "MAIN", "quantity": 1, "allocated_unit_price": 28500},
                {"price_rule_key": "price-1", "sequence": 2, "item_code": "OPTION", "quantity": 1, "allocated_unit_price": 6900},
            ],
        )

    def test_set_discount_is_absorbed_by_main_product(self) -> None:
        order = SmartStoreOrder(
            source_row=2,
            order_no="ORDER-1",
            product_order_no="PRODUCT-1",
            paid_at=datetime(2026, 7, 21, 10, 0),
            status="구매확정",
            product_name="상품 세트",
            options="옵션 핑크",
            quantity=Decimal("1"),
            item_total=Decimal("34900"),
        )
        result = convert_orders([order], self.catalog)
        by_code = {line.item_code: line for line in result.lines}
        self.assertEqual(by_code["MAIN"].unit_price, Decimal("28000.00"))
        self.assertEqual(by_code["OPTION"].unit_price, Decimal("6900.00"))
        self.assertEqual(result.input_total, result.output_total)
        self.assertEqual(result.issues, [])

    def test_export_uses_voucher_date_and_tax_formulas(self) -> None:
        order = SmartStoreOrder(
            source_row=2,
            order_no="ORDER-1",
            product_order_no="PRODUCT-1",
            paid_at=datetime(2026, 7, 21, 10, 0),
            status="구매확정",
            product_name="상품 세트",
            options="옵션 핑크",
            quantity=Decimal("1"),
            item_total=Decimal("34900"),
        )
        result = convert_orders([order], self.catalog)
        with tempfile.TemporaryDirectory() as folder:
            output = Path(folder) / "output.xlsx"
            write_ecount_workbook(output, result, date(2026, 7, 22))
            workbook = load_workbook(output, data_only=False)
            sheet = workbook["이카운트 웹입력"]
            self.assertEqual(sheet["A2"].value, 20260722)
            self.assertEqual(sheet["E2"].value, "00109")
            self.assertEqual(sheet["F2"].value, "300")
            self.assertEqual(sheet["S2"].value, "=ROUND(Q2/1.1*P2,0)")
            self.assertEqual(sheet["T2"].value, "=Q2*P2-S2")
            workbook.close()

    def test_final_amount_is_preferred_and_shipping_is_deduplicated(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "source.xlsx"
            from openpyxl import Workbook

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "배송현황관리"
            sheet.append([
                "상품주문번호", "주문번호", "상품명", "옵션정보", "수량", "결제일", "주문상태",
                "최종 상품별 총 주문금액", "최초 상품별 총 주문금액", "배송비 묶음번호",
                "배송비 합계", "제주/도서 추가배송비", "배송비 할인액",
            ])
            for product_order_no in ("P1", "P2"):
                sheet.append([
                    product_order_no, "ORDER-1", "상품 세트", "옵션 핑크", 1, "2026-07-21", "구매확정",
                    34900, 35900, "SHIP-1", 6000, 3000, 6000,
                ])
            workbook.save(source)
            workbook.close()

            orders = read_smartstore_orders(source, date(2026, 7, 21))
            self.assertEqual([order.item_total for order in orders], [Decimal("34900.00"), Decimal("34900.00")])
            result = convert_orders(orders, self.catalog)
            shipping_lines = [line for line in result.lines if line.is_shipping]
            self.assertEqual(len(result.shipping_charges), 1)
            self.assertEqual(result.shipping_charges[0].effective_amount, Decimal("3000.00"))
            self.assertTrue(result.shipping_charges[0].is_adjusted)
            self.assertEqual(len(shipping_lines), 1)
            self.assertEqual(shipping_lines[0].quantity, Decimal("1"))
            self.assertEqual(shipping_lines[0].unit_price, Decimal("3000.00"))
            self.assertTrue(result.is_reconciled)

    def test_export_is_blocked_when_review_issue_exists(self) -> None:
        order = SmartStoreOrder(
            source_row=2,
            order_no="ORDER-2",
            product_order_no="PRODUCT-2",
            paid_at=datetime(2026, 7, 21, 10, 0),
            status="구매확정",
            product_name="미등록 상품",
            options="",
            quantity=Decimal("1"),
            item_total=Decimal("27800"),
        )
        result = convert_orders([order], self.catalog)
        self.assertEqual(result.amount_difference, Decimal("27800.00"))
        self.assertFalse(result.is_reconciled)
        with tempfile.TemporaryDirectory() as folder:
            with self.assertRaisesRegex(ValueError, "확인 필요 항목"):
                write_ecount_workbook(Path(folder) / "blocked.xlsx", result, date(2026, 7, 22))


if __name__ == "__main__":
    unittest.main()
