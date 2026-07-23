from __future__ import annotations

import sys
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


def reference_path() -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / "assets" / "ecount_item_reference.xlsx"


@lru_cache(maxsize=1)
def load_ecount_items() -> list[dict[str, str | bool]]:
    path = reference_path()
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_index = -1
    code_index = name_index = -1
    for row_number, row in enumerate(rows[:20]):
        normalized = {
            "".join(str(value or "").lower().split()): index
            for index, value in enumerate(row)
        }
        if "품목코드" in normalized and "품명및규격" in normalized:
            header_index = row_number
            code_index = normalized["품목코드"]
            name_index = normalized["품명및규격"]
            break
    if header_index < 0:
        raise ValueError("이카운트 품목 기준표에서 '품목코드', '품명 및 규격' 열을 찾지 못했습니다.")

    result: list[dict[str, str | bool]] = []
    seen_codes: set[str] = set()
    for row in rows[header_index + 1 :]:
        code = str(row[code_index] or "").strip() if code_index < len(row) else ""
        name = str(row[name_index] or "").strip() if name_index < len(row) else ""
        if not code or not name or code == "합계":
            continue
        code_key = code.casefold()
        if code_key in seen_codes:
            continue
        seen_codes.add(code_key)
        result.append({
            "item_code": code,
            "standard_name": name,
            "model": "",
            "color": "",
            "form": "",
            "is_active": True,
        })
    return result
