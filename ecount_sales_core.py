from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_DOWN, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CHANNEL_NAME = "리큐엠_스마트스토어"
MONEY = Decimal("0.01")


def normalize_source(value: str) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", (value or "").lower())


def as_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value).replace(",", "")).quantize(MONEY)
    except (InvalidOperation, ValueError):
        return default


def as_bool(value: Any) -> bool:
    return str(value).strip().lower() not in {"false", "0", "no", "n", ""}


def clean_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_excel_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M:%S", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


@dataclass
class SmartStoreOrder:
    source_row: int
    order_no: str
    product_order_no: str
    paid_at: datetime
    status: str
    product_name: str
    options: str
    quantity: Decimal
    item_total: Decimal
    shipping_bundle_no: str = ""
    shipping_total: Decimal = Decimal("0")
    extra_shipping: Decimal = Decimal("0")
    shipping_discount: Decimal = Decimal("0")

    @property
    def normalized_source(self) -> str:
        return normalize_source(f"{self.product_name}{self.options}")

    @property
    def unit_total(self) -> Decimal:
        if self.quantity == 0:
            return Decimal("0")
        return (self.item_total / self.quantity).quantize(MONEY, rounding=ROUND_HALF_UP)


@dataclass
class VoucherLine:
    customer_code: str
    customer_name: str
    item_code: str
    item_name: str
    quantity: Decimal
    unit_price: Decimal
    warehouse: str
    source_count: int = 1
    source_orders: list[str] = field(default_factory=list)
    is_shipping: bool = False
    needs_review: bool = False
    review_reason: str = ""

    @property
    def total(self) -> Decimal:
        return (self.quantity * self.unit_price).quantize(MONEY, rounding=ROUND_HALF_UP)


@dataclass
class ReviewIssue:
    source_row: int
    order_no: str
    product_name: str
    options: str
    quantity: Decimal
    amount: Decimal
    reason: str


@dataclass
class ShippingCharge:
    bundle_key: str
    order_no: str
    source_row: int
    shipping_total: Decimal
    extra_shipping: Decimal
    shipping_discount: Decimal
    effective_amount: Decimal

    @property
    def is_adjusted(self) -> bool:
        return self.extra_shipping != 0 or self.shipping_discount != 0


@dataclass
class ConversionResult:
    orders: list[SmartStoreOrder]
    lines: list[VoucherLine]
    issues: list[ReviewIssue]
    shipping_charges: list[ShippingCharge] = field(default_factory=list)

    @property
    def input_total(self) -> Decimal:
        return sum((row.item_total for row in self.orders), Decimal("0"))

    @property
    def output_total(self) -> Decimal:
        return sum((row.total for row in self.lines), Decimal("0"))

    @property
    def shipping_total(self) -> Decimal:
        return sum((row.effective_amount for row in self.shipping_charges), Decimal("0"))

    @property
    def excluded_total(self) -> Decimal:
        return sum((row.amount for row in self.issues if "배송비" not in row.reason), Decimal("0"))

    @property
    def expected_output_total(self) -> Decimal:
        return self.input_total + self.shipping_total

    @property
    def amount_difference(self) -> Decimal:
        return self.expected_output_total - self.output_total

    @property
    def is_reconciled(self) -> bool:
        return self.amount_difference == 0


class ReferenceCatalog:
    def __init__(
        self,
        items: Iterable[dict[str, Any]],
        channels: Iterable[dict[str, Any]],
        mappings: Iterable[dict[str, Any]],
        mapping_components: Iterable[dict[str, Any]],
        price_rules: Iterable[dict[str, Any]],
        price_components: Iterable[dict[str, Any]],
    ) -> None:
        self.items = {str(row.get("item_code", "")): row for row in items if row.get("item_code")}
        self.channels = {
            str(row.get("source_name", "")): row
            for row in channels
            if row.get("source_name") and as_bool(row.get("is_active", True))
        }
        components_by_mapping: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in mapping_components:
            components_by_mapping[str(row.get("mapping_key", ""))].append(row)
        self.mappings: dict[tuple[str, str], dict[str, Any]] = {}
        for row in mappings:
            if not as_bool(row.get("is_active", True)) or row.get("review_status") != "confirmed":
                continue
            entry = dict(row)
            entry["components"] = sorted(
                components_by_mapping.get(str(row.get("mapping_key", "")), []),
                key=lambda x: int(x.get("sequence", 0)),
            )
            self.mappings[(str(row.get("source_channel", "")), str(row.get("normalized_source", "")))] = entry

        components_by_rule: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in price_components:
            components_by_rule[str(row.get("price_rule_key", ""))].append(row)
        self.price_templates: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
        for row in price_rules:
            if not as_bool(row.get("is_active", True)):
                continue
            if row.get("review_status") not in {"confirmed", "amount_mismatch"}:
                continue
            entry = dict(row)
            entry["total_unit_price"] = as_decimal(row.get("total_unit_price"))
            entry["components"] = sorted(
                components_by_rule.get(str(row.get("price_rule_key", "")), []),
                key=lambda x: int(x.get("sequence", 0)),
            )
            if entry["components"] and all(component.get("item_code") for component in entry["components"]):
                product_key = normalize_source(str(row.get("source_product_name", "")))
                option_key = normalize_source(str(row.get("source_options", "")))
                canonical_key = product_key if product_key == option_key else f"{product_key}{option_key}"
                lookup_keys = {str(row.get("normalized_source", "")), canonical_key, product_key, option_key}
                for lookup_key in lookup_keys:
                    if lookup_key:
                        self.price_templates[(str(row.get("source_channel", "")), lookup_key)].append(entry)

    @classmethod
    def from_csv_dir(cls, folder: str | Path) -> "ReferenceCatalog":
        folder = Path(folder)

        def read(name: str) -> list[dict[str, Any]]:
            with (folder / name).open("r", encoding="utf-8-sig", newline="") as handle:
                return list(csv.DictReader(handle))

        return cls(
            read("ecount_item_reference.csv"),
            read("ecount_sales_channels.csv"),
            read("ecount_product_mappings.csv"),
            read("ecount_product_mapping_components.csv"),
            read("ecount_price_rules.csv"),
            read("ecount_price_rule_components.csv"),
        )


def read_smartstore_orders(path: str | Path, target_date: date) -> list[SmartStoreOrder]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["구매확정내역"] if "구매확정내역" in workbook.sheetnames else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header: list[str] | None = None
        header_row = 0
        for row_number, row in enumerate(rows, start=1):
            names = [str(value or "").strip() for value in row]
            required = {"주문번호", "상품명", "옵션정보", "수량", "결제일"}
            if required.issubset(set(names)):
                header = names
                header_row = row_number
                break
            if row_number >= 30:
                break
        if header is None:
            raise ValueError("스마트스토어 헤더(주문번호/상품명/옵션정보/수량/결제일)를 찾지 못했습니다.")
        indexes = {name: index for index, name in enumerate(header)}
        result: list[SmartStoreOrder] = []
        for source_row, row in enumerate(rows, start=header_row + 1):
            if not any(value not in (None, "") for value in row):
                continue
            paid_at = parse_excel_date(row[indexes["결제일"]])
            if paid_at is None or paid_at.date() != target_date:
                continue
            quantity = as_decimal(row[indexes["수량"]])
            if quantity <= 0:
                continue
            amount = Decimal("0")
            for field_name in ("최종 상품별 총 주문금액", "최초 상품별 총 주문금액"):
                if field_name in indexes and row[indexes[field_name]] not in (None, ""):
                    amount = as_decimal(row[indexes[field_name]])
                    break
            if amount == 0 and "상품가격" in indexes:
                item_price = as_decimal(row[indexes["상품가격"]])
                option_price = as_decimal(row[indexes.get("옵션가격", indexes["상품가격"])] if "옵션가격" in indexes else 0)
                discount = as_decimal(row[indexes.get("최종 상품별 할인액", indexes["상품가격"])] if "최종 상품별 할인액" in indexes else 0)
                amount = ((item_price + option_price) * quantity - discount).quantize(MONEY)
            result.append(
                SmartStoreOrder(
                    source_row=source_row,
                    order_no=clean_identifier(row[indexes["주문번호"]]),
                    product_order_no=clean_identifier(row[indexes.get("상품주문번호", indexes["주문번호"])]),
                    paid_at=paid_at,
                    status=str(row[indexes.get("주문상태", indexes["주문번호"])] or ""),
                    product_name=str(row[indexes["상품명"]] or "").strip(),
                    options=str(row[indexes["옵션정보"]] or "").strip(),
                    quantity=quantity,
                    item_total=amount,
                    shipping_bundle_no=clean_identifier(row[indexes["배송비 묶음번호"]]) if "배송비 묶음번호" in indexes else "",
                    shipping_total=as_decimal(row[indexes["배송비 합계"]]) if "배송비 합계" in indexes else Decimal("0"),
                    extra_shipping=as_decimal(row[indexes["제주/도서 추가배송비"]]) if "제주/도서 추가배송비" in indexes else Decimal("0"),
                    shipping_discount=as_decimal(row[indexes["배송비 할인액"]]) if "배송비 할인액" in indexes else Decimal("0"),
                )
            )
        return result
    finally:
        workbook.close()


def convert_orders(
    orders: list[SmartStoreOrder],
    catalog: ReferenceCatalog,
    channel_name: str = CHANNEL_NAME,
    default_warehouse: str = "300",
) -> ConversionResult:
    channel = catalog.channels.get(channel_name, {})
    customer_code = str(channel.get("ecount_customer_code") or "AC008712")
    customer_name = str(channel.get("ecount_customer_name") or "샵N")
    raw_lines: list[VoucherLine] = []
    issues: list[ReviewIssue] = []
    shipping_charges = _collect_shipping_charges(orders, issues)

    for order in orders:
        if any(word in order.status for word in ("취소", "반품", "교환")):
            issues.append(_issue(order, f"주문상태 확인 필요: {order.status}"))
            continue
        mapping = catalog.mappings.get((channel_name, order.normalized_source))
        if not mapping:
            reason = "상품/옵션 조합이 DB에 없습니다."
            issues.append(_issue(order, reason))
            _append_review_line(raw_lines, customer_code, customer_name, order, default_warehouse, reason)
            continue
        components = mapping.get("components", [])
        if mapping.get("mapping_type") == "single":
            if len(components) != 1:
                reason = "단품 매핑의 구성품 수가 1개가 아닙니다."
                issues.append(_issue(order, reason))
                _append_review_line(raw_lines, customer_code, customer_name, order, default_warehouse, reason)
                continue
            component = components[0]
            _append_exact_total_lines(
                raw_lines,
                catalog,
                customer_code,
                customer_name,
                component,
                order.quantity,
                order.item_total,
                default_warehouse,
                order.order_no,
            )
            continue

        templates = catalog.price_templates.get((channel_name, order.normalized_source), [])
        if not templates:
            reason = "세트 가격 배분 기준이 DB에 없습니다."
            issues.append(_issue(order, reason))
            _append_review_line(raw_lines, customer_code, customer_name, order, default_warehouse, reason)
            continue
        template = min(templates, key=lambda row: abs(row["total_unit_price"] - order.unit_total))
        price_components = template["components"]
        if len(price_components) < 2:
            reason = "세트 가격 배분 구성품이 부족합니다."
            issues.append(_issue(order, reason))
            _append_review_line(raw_lines, customer_code, customer_name, order, default_warehouse, reason)
            continue
        fixed_total = sum(
            (as_decimal(row.get("allocated_unit_price")) * as_decimal(row.get("quantity"), Decimal("1")))
            for row in price_components[1:]
        )
        main = price_components[0]
        main_quantity = as_decimal(main.get("quantity"), Decimal("1"))
        main_price = ((order.unit_total - fixed_total) / main_quantity).quantize(MONEY, rounding=ROUND_HALF_UP)
        if main_price < 0:
            reason = f"세트 차감 후 본품 단가가 음수입니다: {main_price}"
            issues.append(_issue(order, reason))
            _append_review_line(raw_lines, customer_code, customer_name, order, default_warehouse, reason)
            continue
        _append_line(raw_lines, catalog, customer_code, customer_name, main, order.quantity * main_quantity, main_price, default_warehouse, order.order_no)
        for component in price_components[1:]:
            component_quantity = as_decimal(component.get("quantity"), Decimal("1"))
            _append_line(
                raw_lines,
                catalog,
                customer_code,
                customer_name,
                component,
                order.quantity * component_quantity,
                as_decimal(component.get("allocated_unit_price")),
                default_warehouse,
                order.order_no,
            )

    shipping_item = catalog.items.get("택배운송비", {})
    shipping_name = str(
        shipping_item.get("representative_name")
        or shipping_item.get("item_name")
        or shipping_item.get("standard_name")
        or "배송비"
    )
    shipping_counts: dict[Decimal, int] = defaultdict(int)
    shipping_orders: dict[Decimal, list[str]] = defaultdict(list)
    for charge in shipping_charges:
        if charge.effective_amount > 0:
            shipping_counts[charge.effective_amount] += 1
            shipping_orders[charge.effective_amount].append(charge.order_no)
    for unit_price, count in shipping_counts.items():
        raw_lines.append(
            VoucherLine(
                customer_code=customer_code,
                customer_name=customer_name,
                item_code="택배운송비",
                item_name=shipping_name,
                quantity=Decimal(count),
                unit_price=unit_price,
                warehouse=default_warehouse,
                source_count=count,
                source_orders=shipping_orders[unit_price],
                is_shipping=True,
            )
        )

    aggregated: dict[tuple[Any, ...], VoucherLine] = {}
    for line in raw_lines:
        key = (
            line.customer_code,
            line.item_code,
            line.item_name if line.needs_review else "",
            line.warehouse,
            line.unit_price,
            line.source_orders[0] if line.needs_review and line.source_orders else "",
        )
        if key not in aggregated:
            aggregated[key] = line
        else:
            current = aggregated[key]
            current.quantity += line.quantity
            current.source_count += line.source_count
            current.source_orders.extend(line.source_orders)
    lines = sorted(aggregated.values(), key=lambda row: (row.item_code, row.unit_price, row.warehouse))
    return ConversionResult(orders=orders, lines=lines, issues=issues, shipping_charges=shipping_charges)


def _collect_shipping_charges(
    orders: list[SmartStoreOrder],
    issues: list[ReviewIssue],
) -> list[ShippingCharge]:
    grouped: dict[str, list[SmartStoreOrder]] = defaultdict(list)
    for order in orders:
        key = order.shipping_bundle_no or order.order_no
        if key:
            grouped[key].append(order)

    charges: list[ShippingCharge] = []
    for key, group in grouped.items():
        signatures = {
            (order.shipping_total, order.extra_shipping, order.shipping_discount)
            for order in group
        }
        first = group[0]
        if len(signatures) != 1:
            issues.append(_issue(first, "같은 배송비 묶음번호에 서로 다른 배송비가 있습니다."))
            continue
        shipping_total, extra_shipping, shipping_discount = next(iter(signatures))
        effective = shipping_total + extra_shipping
        if effective < 0:
            issues.append(_issue(first, f"배송비 합계와 추가배송비의 합이 음수입니다: {effective}"))
            continue
        charges.append(
            ShippingCharge(
                bundle_key=key,
                order_no=first.order_no,
                source_row=first.source_row,
                shipping_total=shipping_total,
                extra_shipping=extra_shipping,
                shipping_discount=shipping_discount,
                effective_amount=effective,
            )
        )
    return charges


def _append_line(
    target: list[VoucherLine],
    catalog: ReferenceCatalog,
    customer_code: str,
    customer_name: str,
    component: dict[str, Any],
    quantity: Decimal,
    unit_price: Decimal,
    default_warehouse: str,
    order_no: str,
) -> None:
    item_code = str(component.get("item_code") or "")
    item = catalog.items.get(item_code, {})
    item_name = str(item.get("representative_name") or item.get("item_name") or item.get("standard_name") or item_code)
    warehouse = "100" if "QM4100" in f"{item_code} {item_name}".upper() else default_warehouse
    target.append(
        VoucherLine(
            customer_code=customer_code,
            customer_name=customer_name,
            item_code=item_code,
            item_name=item_name,
            quantity=quantity,
            unit_price=unit_price.quantize(MONEY, rounding=ROUND_HALF_UP),
            warehouse=warehouse,
            source_orders=[order_no],
        )
    )


def _append_exact_total_lines(
    target: list[VoucherLine],
    catalog: ReferenceCatalog,
    customer_code: str,
    customer_name: str,
    component: dict[str, Any],
    quantity: Decimal,
    total: Decimal,
    default_warehouse: str,
    order_no: str,
) -> None:
    if (
        quantity > 0
        and quantity == quantity.to_integral_value()
        and total == total.to_integral_value()
    ):
        base_price = (total / quantity).quantize(Decimal("1"), rounding=ROUND_DOWN)
        higher_price_quantity = total - (base_price * quantity)
        base_price_quantity = quantity - higher_price_quantity
        if base_price_quantity > 0:
            _append_line(
                target, catalog, customer_code, customer_name, component,
                base_price_quantity, base_price, default_warehouse, order_no,
            )
        if higher_price_quantity > 0:
            _append_line(
                target, catalog, customer_code, customer_name, component,
                higher_price_quantity, base_price + 1, default_warehouse, order_no,
            )
        return
    _append_line(
        target, catalog, customer_code, customer_name, component,
        quantity, (total / quantity).quantize(MONEY, rounding=ROUND_HALF_UP),
        default_warehouse, order_no,
    )


def _append_review_line(
    target: list[VoucherLine],
    customer_code: str,
    customer_name: str,
    order: SmartStoreOrder,
    default_warehouse: str,
    reason: str,
) -> None:
    splits = [(order.quantity, order.unit_total)]
    if (
        order.quantity > 0
        and order.quantity == order.quantity.to_integral_value()
        and order.item_total == order.item_total.to_integral_value()
    ):
        base_price = (order.item_total / order.quantity).quantize(Decimal("1"), rounding=ROUND_DOWN)
        higher_price_quantity = order.item_total - (base_price * order.quantity)
        splits = [
            (quantity, unit_price)
            for quantity, unit_price in (
                (order.quantity - higher_price_quantity, base_price),
                (higher_price_quantity, base_price + 1),
            )
            if quantity > 0
        ]
    for quantity, unit_price in splits:
        target.append(
            VoucherLine(
                customer_code=customer_code,
                customer_name=customer_name,
                item_code="",
                item_name=f"[확인필요] {order.product_name} {order.options}".strip(),
                quantity=quantity,
                unit_price=unit_price,
                warehouse=default_warehouse,
                source_orders=[order.order_no],
                needs_review=True,
                review_reason=reason,
            )
        )


def _issue(order: SmartStoreOrder, reason: str) -> ReviewIssue:
    return ReviewIssue(order.source_row, order.order_no, order.product_name, order.options, order.quantity, order.item_total, reason)


def write_ecount_workbook(
    path: str | Path,
    result: ConversionResult,
    voucher_date: date,
    manager_code: str = "00109",
) -> None:
    if not result.is_reconciled:
        raise ValueError(f"금액 차이 {result.amount_difference:,.0f}원이 있어 전표를 저장할 수 없습니다.")
    workbook = Workbook()
    upload = workbook.active
    upload.title = "이카운트 웹입력"
    headers = [
        "일자", "순번", "거래처코드", "거래처명", "담당자", "출하창고", "거래유형", "통화", "환율", "계좌번호", "미수금",
        "특이사항", "품목코드", "품목명", "규격", "수량", "단가", "외화금액", "공급가액", "부가세", "비고", "생산전표생성",
    ]
    upload.append(headers)
    date_number = int(voucher_date.strftime("%Y%m%d"))
    for index, line in enumerate(result.lines, start=2):
        upload.append([
            date_number, None, line.customer_code, line.customer_name, manager_code, line.warehouse, None, None, None, None, None, None,
            line.item_code, line.item_name, None, float(line.quantity), float(line.unit_price), None,
            f"=ROUND(Q{index}/1.1*P{index},0)", f"=Q{index}*P{index}-S{index}",
            f"확인필요: {line.review_reason}" if line.needs_review else None,
            None,
        ])
    _style_upload_sheet(upload, result.lines)

    review = workbook.create_sheet("검수결과")
    review.append(["구분", "원본행", "주문번호", "품목/옵션", "수량", "금액", "결과/사유"])
    for issue in result.issues:
        review.append(["확인필요", issue.source_row, issue.order_no, f"{issue.product_name} / {issue.options}", float(issue.quantity), float(issue.amount), issue.reason])
    review.append([])
    review.append(["검수 항목", "결과"])
    review.append(["대상 주문행", len(result.orders)])
    review.append(["자동 변환행", len(result.lines)])
    review.append(["확인 필요행", len(result.issues)])
    review.append(["원본 품목금액", float(result.input_total)])
    review.append(["제외 품목금액", float(result.excluded_total)])
    review.append(["전표 배송비(할인 미차감)", float(result.shipping_total)])
    review.append(["검수 기준금액", float(result.expected_output_total)])
    review.append(["전표 총금액", float(result.output_total)])
    review.append(["금액 차이", float(result.amount_difference)])
    review.append([])
    review.append(["배송비 묶음번호", "원배송비", "추가배송비", "할인액(참고)", "전표 배송비(할인 미차감)", "조정 여부"])
    for charge in result.shipping_charges:
        review.append([
            charge.bundle_key,
            float(charge.shipping_total),
            float(charge.extra_shipping),
            float(charge.shipping_discount),
            float(charge.effective_amount),
            "확인" if charge.is_adjusted else "일반",
        ])
    _style_review_sheet(review)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _style_upload_sheet(sheet: Any, lines: list[VoucherLine]) -> None:
    last_row = len(lines) + 1
    header_fill = PatternFill("solid", fgColor="00A651")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = {1: 11, 3: 14, 4: 14, 5: 10, 6: 10, 13: 25, 14: 24, 16: 10, 17: 14, 19: 16, 20: 14}
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(2, last_row + 1):
        sheet.cell(row, 3).number_format = "@"
        sheet.cell(row, 5).number_format = "@"
        sheet.cell(row, 6).number_format = "@"
        sheet.cell(row, 13).number_format = "@"
        for column in (16, 17, 19, 20):
            sheet.cell(row, column).number_format = "#,##0"
        if lines[row - 2].needs_review:
            for cell in sheet[row]:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            sheet.cell(row, 13).font = Font(color="C00000", bold=True)
            sheet.cell(row, 14).font = Font(color="C00000", bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:V{max(last_row, 1)}"


def _style_review_sheet(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for column, width in {1: 12, 2: 10, 3: 22, 4: 70, 5: 10, 6: 14, 7: 45}.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
    sheet.freeze_panes = "A2"
