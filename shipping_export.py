from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = [
    "주문번호", "판매처", "상품명", "수량", "수령자", "핸드폰", "우편번호",
    "주소", "배송메세지", "송장번호", "일련번호",
]


def export_wekep(orders: list[dict[str, str]], file_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "택배출고"
    sheet.append(HEADERS)
    for order in orders:
        product_name = order.get("matched_product") or order.get("matched_name") or order.get("product_name")
        sheet.append(
            [
                order.get("order_number", ""),
                order.get("channel", ""),
                product_name,
                order.get("quantity", ""),
                order.get("recipient", ""),
                order.get("phone", ""),
                order.get("zipcode", ""),
                order.get("address", ""),
                order.get("message", ""),
                "",
                order.get("serial_number", ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [22, 14, 45, 9, 14, 18, 11, 55, 35, 18, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    workbook.save(Path(file_path))
